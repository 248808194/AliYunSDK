# encoding=utf-8
import pymysql as pms
import openpyxl
import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib


def get_datas(sql, conn):
    cur = conn.cursor()
    cur.execute(sql)
    datas = cur.fetchall()
    cur.close()
    return datas


def get_fields(sql, conn):
    cur = conn.cursor()
    cur.execute(sql)
    fields = cur.description  # 获取所需要的字段名称
    cur.close()
    return fields


def get_excel(data, field, file):
    new = openpyxl.Workbook()  # 激活一个新的sheet
    sheet = new.active
    sheet.title = '背调数据日常指标'  # 给sheet命名
    for col in range(len(field)):  # 将字段名称循环写入excel第一行，因为字段格式列表里包含列表，每个列表的第一元素才是字段名称
        _ = sheet.cell(row=1, column=col + 1,
                       value=u'%s' % field[col][0])  # row代表行数，column代表列数，value代表单元格输入的值，行数和列数都是从1开始，这点于python不同要注意
    for row in range(len(data)):  # 将数据循环写入excel的每个单元格中
        for col in range(len(field)):
            _ = sheet.cell(row=row + 2, column=col + 1,
                           value=u'%s' % data[row][col])  # 因为第一行写了字段名称，所以要从第二行开始写入
    newworkbook = new.save(file)  # 将生成的excel保存，这步是必不可少的
    return newworkbook


def getYesterday():
    # 获取昨天日期的字符串格式的函数
    # 获取今天的日期
    today = datetime.date.today()
    # 获取一天的日期格式数据
    oneday = datetime.timedelta(days=1)
    # 昨天等于今天减去一天
    yesterday = today - oneday
    # 获取昨天日期的格式化字符串
    yesterdaystr = yesterday.strftime('%Y-%m-%d')
    # 返回昨天的字符串
    return yesterdaystr


def query_shebao_order():
    daibanuser_connection = pymysql.connect(host="172.19.143.211", user="reader", password="yie8aa+fec8No5eneeC0",
                                            db="my_ucenter")
    peopleTotal = "select count(*) from my_ucenter.t_user where create_time < curdate()"

    try:
        peopleTotal_result = get_datas(peopleTotal, daibanuser_connection)
        peopleTotal_field = get_fields(peopleTotal, daibanuser_connection)
        aa = str(peopleTotal_result['count(*)'])
        yesterdaystr = getYesterday()
        print(yesterdaystr)
        my_file_name = 'user attribute' + yesterdaystr + '.xlsx'
        file_path = 'C:/Users/Renliwo/Desktop/' + my_file_name
        get_excel(peopleTotal_result, peopleTotal_field, file_path)
        print(peopleTotal_result)
    except Exception as e:
        print('error create category', e)


if __name__ == "__main__":
    query_shebao_order()
